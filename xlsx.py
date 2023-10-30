import openpyxl, smtplib 
from pathlib import Path


file = openpyxl.load_workbook(Path.home() / Path('Member.xlsx')) 
sheets = file.sheetnames
sheet = file['Sheet1']
lastCol = sheet.max_column
latesMonth = sheet.cell(row=1, column=lastCol).value
print(latesMonth)
